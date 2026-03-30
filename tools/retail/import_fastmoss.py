"""
FastMoss TikTok Shop Data Importer

Purpose: Import TikTok Shop sales/GMV data from FastMoss Excel exports.
         Writes ALL shops to tiktok_shop_weekly (time-series) and updates
         enriched_companies for matched shops.
Inputs: FastMoss Excel export file (.xlsx), week date
Outputs: tiktok_shop_weekly rows + enriched_companies updates

Usage:
    # Dry run (show matches without writing)
    python -m tools.retail.import_fastmoss "path/to/export.xlsx" --dry-run

    # Import (auto-detects week as last Monday)
    python -m tools.retail.import_fastmoss "path/to/export.xlsx"

    # Import with explicit week
    python -m tools.retail.import_fastmoss "path/to/export.xlsx" --week 2026-03-23

    # Force overwrite existing enriched_companies data
    python -m tools.retail.import_fastmoss "path/to/export.xlsx" --force
"""

import os
import sys
import argparse
import json
from typing import Dict, Any, Optional, List, Tuple
from datetime import datetime, timezone, date, timedelta

_TOOLS_DIR = os.path.join(os.path.dirname(__file__), "..")
if _TOOLS_DIR not in sys.path:
    sys.path.insert(0, _TOOLS_DIR)


# ---------------------------------------------------------------------------
# Column name mapping — FastMoss exports may use varying column names.
# We normalize by matching against known patterns (case-insensitive).
# ---------------------------------------------------------------------------

COLUMN_PATTERNS = {
    "shop_name": ["nombre de la tienda", "shop name", "store name"],
    "company_name": ["nombre de la empresa", "company name"],
    "shop_type": ["posicionamiento de la tienda", "shop type", "store type"],
    "country": ["país", "pais", "country"],
    "category": ["categoría principal", "categoria principal", "category", "main category"],
    "rating": ["calificación de la tienda", "calificacion de la tienda", "rating", "shop rating"],
    "sales_count": ["ventas [", "ventas", "sales"],
    "gmv": ["ingresos [", "ingresos", "revenue", "gmv"],
    "products": ["de productos activos", "products", "product count", "productos"],
    "influencers": ["número de influencers", "numero de influencers", "influencers"],
    "fastmoss_url": ["enlace a la página", "enlace a la pagina", "fastmoss"],
}


def _normalize_column_name(raw_name: str) -> Optional[str]:
    """Match a raw Excel column header to a canonical field name."""
    raw_lower = raw_name.strip().lower()

    # Skip "comparación mensual" columns — they are MoM % changes, not absolute values
    if "comparaci" in raw_lower and "mensual" in raw_lower:
        return None

    for canonical, patterns in COLUMN_PATTERNS.items():
        for pattern in patterns:
            if pattern == raw_lower or pattern in raw_lower:
                return canonical
    return None


def _parse_numeric(value: Any) -> Optional[float]:
    """
    Parse a potentially formatted number from FastMoss exports.

    Handles formats like:
        52.8mil  → 52800
        MX$11.5millón → 11500000
        $1,234.56 → 1234.56
        -32.32% → -32.32
        1.2K → 1200
        3.5M → 3500000
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s or s == "-" or s.lower() in ("n/a", "null", "none"):
        return None

    # Strip currency prefixes
    import re
    s = re.sub(r'^[A-Z]{0,3}\$', '', s).strip()

    # Strip percentage sign (keep the number)
    s = s.replace("%", "").strip()
    s = s.replace(",", "")

    # Handle Spanish suffixes: "millón"/"millones" → 1M, "mil" → 1K
    multiplier = 1
    if re.search(r'mill[oó]n(es)?$', s, re.IGNORECASE):
        multiplier = 1_000_000
        s = re.sub(r'\s*mill[oó]n(es)?$', '', s, flags=re.IGNORECASE).strip()
    elif s.lower().endswith("mil"):
        multiplier = 1_000
        s = s[:-3].strip()
    elif s.upper().endswith("K"):
        multiplier = 1_000
        s = s[:-1].strip()
    elif s.upper().endswith("M"):
        multiplier = 1_000_000
        s = s[:-1].strip()

    try:
        return float(s) * multiplier
    except ValueError:
        return None


def read_fastmoss_excel(filepath: str) -> Dict[str, Any]:
    """
    Read a FastMoss Excel export and extract shop data.

    Returns:
        Dict with:
            - success: bool
            - data: {shops: list of dicts, columns_found: dict, total_rows: int}
            - error: str or None
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # Read header row
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return {"success": False, "data": {}, "error": "Empty file or no header row"}

        # Map columns
        column_map = {}  # index -> canonical name
        raw_headers = []
        for i, cell in enumerate(header_row):
            raw = str(cell).strip() if cell else ""
            raw_headers.append(raw)
            canonical = _normalize_column_name(raw)
            if canonical:
                column_map[i] = canonical

        if "shop_name" not in column_map.values():
            return {
                "success": False,
                "data": {"raw_headers": raw_headers},
                "error": f"Could not find shop name column. Headers found: {raw_headers}",
            }

        # Read data rows
        shops = []
        for row in rows:
            if not row or all(c is None for c in row):
                continue

            shop = {}
            TEXT_FIELDS = {"shop_name", "company_name", "shop_type", "country",
                           "category", "fastmoss_url"}
            for i, canonical in column_map.items():
                if i < len(row):
                    raw_val = row[i]
                    if canonical in TEXT_FIELDS:
                        shop[canonical] = str(raw_val).strip() if raw_val else None
                    else:
                        shop[canonical] = _parse_numeric(raw_val)

            if shop.get("shop_name"):
                shops.append(shop)

        wb.close()

        columns_found = {v: raw_headers[k] for k, v in column_map.items()}

        return {
            "success": True,
            "data": {
                "shops": shops,
                "columns_found": columns_found,
                "total_rows": len(shops),
            },
            "error": None,
        }

    except Exception as e:
        return {"success": False, "data": {}, "error": f"Failed to read Excel: {str(e)}"}


def _normalize_for_match(name: str) -> str:
    """Normalize a brand/shop name for matching by removing common suffixes and noise."""
    import re
    s = name.lower().strip()
    # Remove common suffixes: MX, Mexico, México, Tienda, Shop, Official, Oficial, SA DE CV
    s = re.sub(r'\b(mx|mexico|méxico|tienda|shop|official|oficial|store)\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bsa\s+de\s+cv\b', '', s, flags=re.IGNORECASE)
    # Remove trailing/leading dashes, dots, spaces
    s = re.sub(r'[^a-záéíóúñü0-9]+', ' ', s).strip()
    return s


def match_shops_to_companies(
    shops: List[Dict],
    supabase_client: Any,
) -> List[Dict]:
    """
    Match FastMoss shop names to enriched_companies using multi-strategy matching.

    Strategies (in order of confidence):
    1. Exact match on company_name
    2. Normalized name match (removing MX, Mexico, Shop, etc.)
    3. Domain-based match (domain without TLD contained in shop name)
    4. Fuzzy match (token_set_ratio >= 85)

    Returns list of dicts: {shop: original shop dict, domain: str, brand_name: str, score: float}
    """
    from rapidfuzz.fuzz import token_set_ratio

    # Fetch all MEX companies from Supabase
    rows = supabase_client.select(
        "enriched_companies",
        columns="domain,company_name,geography",
        eq={"geography": "MEX"},
    )

    if not rows:
        return []

    # Pre-compute normalized names and domain bases for all companies
    company_index = []
    for row in rows:
        company = row.get("company_name", "") or ""
        domain = row.get("domain", "") or ""
        # Extract meaningful domain base
        domain_parts = domain.lower().split(".")
        domain_base = domain_parts[0]
        if domain_base in ("www", "tienda", "shop", "store", "mx") and len(domain_parts) > 2:
            domain_base = domain_parts[1]

        company_index.append({
            "row": row,
            "company_lower": company.lower().strip(),
            "company_normalized": _normalize_for_match(company),
            "domain_base": domain_base,
        })

    matches = []
    unmatched = []

    for shop in shops:
        shop_name = shop.get("shop_name", "")
        fm_company = shop.get("company_name", "")
        if not shop_name and not fm_company:
            continue

        # Build search variants from FastMoss data
        search_names = [n.lower().strip() for n in [shop_name, fm_company] if n]
        search_normalized = [_normalize_for_match(n) for n in [shop_name, fm_company] if n]

        best_match = None
        best_score = 0

        for ci in company_index:
            row = ci["row"]
            company_lower = ci["company_lower"]
            company_norm = ci["company_normalized"]
            domain_base = ci["domain_base"]

            for idx, search_name in enumerate(search_names):
                search_norm = search_normalized[idx] if idx < len(search_normalized) else search_name

                # Strategy 1: Exact match on company name
                if search_name == company_lower:
                    best_match = row
                    best_score = 100
                    break

                # Strategy 2: Normalized name match
                if (search_norm and company_norm and len(company_norm) >= 4
                        and (search_norm == company_norm
                             or (company_norm in search_norm and
                                 len(company_norm) / len(search_norm) >= 0.6)
                             or (search_norm in company_norm and
                                 len(search_norm) / len(company_norm) >= 0.6))):
                    score = 95
                    if score > best_score:
                        best_score = score
                        best_match = row

                # Strategy 3: Domain-based match
                if domain_base and len(domain_base) >= 5:
                    # Check if domain base appears as a substring in normalized shop name
                    if domain_base in search_norm and len(domain_base) / len(search_norm) >= 0.4:
                        score = 90
                        if score > best_score:
                            best_score = score
                            best_match = row

                # Strategy 4: Fuzzy match
                if company_lower and len(company_lower) >= 4:
                    score = token_set_ratio(search_name, company_lower)
                    if score > best_score and score >= 85:
                        best_score = score
                        best_match = row

            if best_score == 100:
                break

        if best_match:
            matches.append({
                "shop": shop,
                "domain": best_match["domain"],
                "brand_name": best_match.get("company_name", ""),
                "score": best_score,
            })
        else:
            unmatched.append(shop_name or fm_company)

    if unmatched:
        print(f"\n  [WARN] {len(unmatched)} shops could not be matched:")
        for name in unmatched[:10]:
            print(f"    - {name}")
        if len(unmatched) > 10:
            print(f"    ... and {len(unmatched) - 10} more")

    return matches


def write_tiktok_shop_data(
    matches: List[Dict],
    supabase_client: Any,
    force: bool = False,
) -> Dict[str, Any]:
    """
    Write TikTok Shop metrics from FastMoss to Supabase.

    Returns summary dict with counts.
    """
    import requests

    written = 0
    skipped = 0
    errors = 0

    for match in matches:
        domain = match["domain"]
        shop = match["shop"]

        try:
            # Check if already has tiktok_shop data (unless force)
            if not force:
                existing = supabase_client.select(
                    "enriched_companies",
                    columns="domain,tiktok_shop_last_synced",
                    eq={"domain": domain},
                    limit=1,
                )
                if existing and existing[0].get("tiktok_shop_last_synced"):
                    skipped += 1
                    continue

            row = {
                "on_tiktok_shop": True,
            }

            # Map FastMoss fields to Supabase columns
            if shop.get("gmv") is not None:
                row["tiktok_shop_gmv"] = shop["gmv"]
            if shop.get("products") is not None:
                row["tiktok_shop_products"] = int(shop["products"])
            if shop.get("rating") is not None:
                row["tiktok_shop_rating"] = shop["rating"]
            if shop.get("sales_count") is not None:
                row["tiktok_shop_sales_estimate"] = shop["sales_count"]
            row["tiktok_shop_last_synced"] = datetime.now(timezone.utc).isoformat()

            resp = requests.patch(
                f"{supabase_client.rest_url}/enriched_companies",
                headers=supabase_client.headers,
                params={"domain": f"eq.{domain}"},
                json=row,
                timeout=15,
            )

            # If new columns don't exist yet, retry with only on_tiktok_shop
            if resp.status_code == 400 and "does not exist" in resp.text:
                row_minimal = {
                    "on_tiktok_shop": True,
                }
                resp = requests.patch(
                    f"{supabase_client.rest_url}/enriched_companies",
                    headers=supabase_client.headers,
                    params={"domain": f"eq.{domain}"},
                    json=row_minimal,
                    timeout=15,
                )

            resp.raise_for_status()
            written += 1
            print(f"  [OK] {domain} <- {shop.get('shop_name', '?')} "
                  f"(GMV: {shop.get('gmv', '?')}, products: {shop.get('products', '?')})")

        except Exception as e:
            errors += 1
            print(f"  [ERR] {domain}: {e}")

    return {"written": written, "skipped": skipped, "errors": errors}


def _get_week_start(date_str: Optional[str] = None) -> str:
    """Return the Monday date string for the given week. Defaults to last Monday."""
    if date_str:
        d = date.fromisoformat(date_str)
    else:
        d = date.today()
    # Go back to Monday of this week
    d = d - timedelta(days=d.weekday())
    return d.isoformat()


def write_tiktok_weekly_data(
    shops: List[Dict],
    matches: List[Dict],
    week_start: str,
    supabase_client: Any,
) -> Dict[str, Any]:
    """
    Write ALL shops to tiktok_shop_weekly table (upsert on shop_name+week_start).

    For matched shops, sets matched_domain. For unmatched shops, matched_domain is null.
    Returns summary counts.
    """
    import requests

    # Build a lookup from shop_name -> matched domain
    match_lookup = {}
    for m in matches:
        shop_name = m["shop"].get("shop_name", "")
        if shop_name:
            match_lookup[shop_name] = m["domain"]

    written = 0
    errors = 0

    # Batch upsert in chunks of 50
    batch = []
    for shop in shops:
        shop_name = shop.get("shop_name")
        if not shop_name:
            continue

        row = {
            "shop_name": shop_name,
            "company_name": shop.get("company_name"),
            "shop_type": shop.get("shop_type"),
            "country": shop.get("country", "MX"),
            "category": shop.get("category"),
            "rating": shop.get("rating"),
            "sales_count": shop.get("sales_count"),
            "gmv": shop.get("gmv"),
            "products": int(shop["products"]) if shop.get("products") is not None else None,
            "influencers": int(shop["influencers"]) if shop.get("influencers") is not None else None,
            "fastmoss_url": (shop.get("fastmoss_url") or "").replace("/zh/", "/en/") or None,
            "week_start": week_start,
            "matched_domain": match_lookup.get(shop_name),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        batch.append(row)

        if len(batch) >= 50:
            result = _upsert_weekly_batch(batch, supabase_client)
            written += result["written"]
            errors += result["errors"]
            batch = []

    # Flush remaining
    if batch:
        result = _upsert_weekly_batch(batch, supabase_client)
        written += result["written"]
        errors += result["errors"]

    return {"written": written, "errors": errors}


def _upsert_weekly_batch(batch: List[Dict], supabase_client: Any) -> Dict[str, int]:
    """Upsert a batch of rows to tiktok_shop_weekly via PostgREST."""
    import requests

    headers = {**supabase_client.headers}
    headers["Prefer"] = "resolution=merge-duplicates"

    try:
        resp = requests.post(
            f"{supabase_client.rest_url}/tiktok_shop_weekly",
            headers=headers,
            params={"on_conflict": "shop_name,week_start"},
            json=batch,
            timeout=30,
        )
        if resp.status_code in (200, 201):
            return {"written": len(batch), "errors": 0}
        else:
            print(f"  [ERR] Batch upsert failed: {resp.status_code} {resp.text[:200]}")
            return {"written": 0, "errors": len(batch)}
    except Exception as e:
        print(f"  [ERR] Batch upsert exception: {e}")
        return {"written": 0, "errors": len(batch)}


def import_fastmoss(
    filepath: str,
    dry_run: bool = False,
    force: bool = False,
    week: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Main entry point: read FastMoss Excel, match to companies, write to Supabase.

    Writes ALL shops to tiktok_shop_weekly (time-series), and updates
    enriched_companies only for matched shops.

    Args:
        filepath: Path to FastMoss Excel export
        dry_run: If True, only show matches without writing
        force: If True, overwrite existing enriched_companies data
        week: Week start date (YYYY-MM-DD, must be Monday). Defaults to last Monday.

    Returns:
        Dict with success, data, error
    """
    week_start = _get_week_start(week)
    print(f"Reading FastMoss export: {filepath}")
    print(f"  Week: {week_start}")

    # Step 1: Read Excel
    read_result = read_fastmoss_excel(filepath)
    if not read_result["success"]:
        return read_result

    shops = read_result["data"]["shops"]
    columns = read_result["data"]["columns_found"]
    print(f"  Found {len(shops)} shops")
    print(f"  Columns mapped: {json.dumps(columns, ensure_ascii=False)}")

    if not shops:
        return {"success": False, "data": {}, "error": "No shop data found in file"}

    # Step 2: Connect to Supabase and match
    try:
        from export.supabase_writer import get_client
        client = get_client()
    except Exception as e:
        return {"success": False, "data": {}, "error": f"Supabase connection failed: {e}"}

    print(f"\nMatching shops to enriched_companies...")
    matches = match_shops_to_companies(shops, client)
    print(f"  Matched: {len(matches)}/{len(shops)}")

    # Show matches if any
    if matches:
        print(f"\n  {'FastMoss Shop':<35} {'Score':>5}  {'Domain':<30} {'Company'}")
        print(f"  {'-' * 35} {'-' * 5}  {'-' * 30} {'-' * 20}")
        for m in matches:
            shop_name = m["shop"].get("shop_name", "?")[:35]
            print(f"  {shop_name:<35} {m['score']:>5.0f}  {m['domain']:<30} {m['brand_name']}")

    if dry_run:
        print(f"\n  [DRY RUN] No data written. Use without --dry-run to import.")
        return {
            "success": True,
            "data": {
                "matched": len(matches),
                "total_shops": len(shops),
                "week_start": week_start,
                "dry_run": True,
            },
            "error": None,
        }

    # Step 3: Write ALL shops to tiktok_shop_weekly
    print(f"\nWriting {len(shops)} shops to tiktok_shop_weekly (week: {week_start})...")
    weekly_result = write_tiktok_weekly_data(shops, matches, week_start, client)
    print(f"  Weekly: {weekly_result['written']} written, {weekly_result['errors']} errors")

    # Step 4: Update enriched_companies for matched shops only
    enriched_result = {"written": 0, "skipped": 0, "errors": 0}
    if matches:
        print(f"\nUpdating enriched_companies for {len(matches)} matched shops...")
        enriched_result = write_tiktok_shop_data(matches, client, force=force)
        print(f"  Enriched: {enriched_result['written']} written, "
              f"{enriched_result['skipped']} skipped, {enriched_result['errors']} errors")

    total_errors = weekly_result["errors"] + enriched_result["errors"]
    print(f"\n  Done. Week: {week_start}, "
          f"shops: {weekly_result['written']}, "
          f"matched: {len(matches)}, "
          f"enriched: {enriched_result['written']}")

    return {
        "success": total_errors == 0,
        "data": {
            "total_shops": len(shops),
            "matched": len(matches),
            "week_start": week_start,
            "weekly_written": weekly_result["written"],
            "enriched_written": enriched_result["written"],
        },
        "error": None,
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import FastMoss TikTok Shop data")
    parser.add_argument("filepath", help="Path to FastMoss Excel export (.xlsx)")
    parser.add_argument("--dry-run", action="store_true", help="Show matches without writing")
    parser.add_argument("--force", action="store_true", help="Overwrite existing enriched_companies data")
    parser.add_argument("--week", help="Week start date (YYYY-MM-DD, must be Monday). Defaults to last Monday.")
    args = parser.parse_args()

    print("FastMoss TikTok Shop Importer")
    print("=" * 60)

    result = import_fastmoss(args.filepath, dry_run=args.dry_run, force=args.force, week=args.week)

    if not result["success"]:
        print(f"\n[ERROR] {result.get('error', 'Unknown error')}")
        sys.exit(1)
